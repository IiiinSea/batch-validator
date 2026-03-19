#!/usr/bin/env python3
"""
Excel utility functions for reading and writing validation results.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any


def read_excel_data(excel_path: str) -> pd.DataFrame:
    """
    Read Excel file into a pandas DataFrame.

    Args:
        excel_path: Path to the Excel file

    Returns:
        DataFrame containing the Excel data
    """
    return pd.read_excel(excel_path)


def add_validation_columns(
    df: pd.DataFrame, validation_results: List[Dict[str, Any]]
) -> pd.DataFrame:
    """
    Add validation result columns to the DataFrame.

    Args:
        df: Original DataFrame
        validation_results: List of validation results for each row
            Each result is a dict with keys: 'row_index' and validation fields

    Returns:
        DataFrame with added validation columns
    """
    # Define validation column names
    validation_columns = [
        "校验1-发布平台",
        "校验2-文章标题",
        "校验3-发布时间",
        "校验4-发布账号",
        "校验5-粉丝数量",
        "校验6-阅读量",
        "校验7-见刊位置",
    ]

    # Initialize validation columns if they don't exist
    for col in validation_columns:
        if col not in df.columns:
            df[col] = ""

    # Fill in validation results
    for result in validation_results:
        row_idx = result["row_index"]
        for i, col in enumerate(validation_columns, start=1):
            key = f"validation_{i}"
            if key in result:
                df.at[row_idx, col] = result[key]

    return df


def save_excel_with_validation(
    df: pd.DataFrame, output_path: str, original_path: str = None
):
    """
    Save DataFrame to Excel file, preserving formatting if possible.

    Args:
        df: DataFrame to save
        output_path: Path for the output Excel file
        original_path: Optional path to original file to preserve formatting
    """
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    output_path = Path(output_path)

    if original_path:
        # Load original workbook to preserve formatting
        wb = load_workbook(original_path)
        ws = wb.active

        # Update existing data and add new columns
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)
    else:
        # Simple save without formatting preservation
        df.to_excel(output_path, index=False, engine="openpyxl")

    print(f"Excel file saved to: {output_path}")


def create_validation_result(
    row_index: int,
    platform_match: str,
    title_match: str,
    time_match: str,
    account_match: str,
    follower_check: str,
    view_check: str,
    position_check: str,
) -> Dict[str, Any]:
    """
    Create a validation result dictionary.

    Args:
        row_index: Row index in the DataFrame
        platform_match: "是" or "否：原因"
        title_match: "是" or "否：原因"
        time_match: "是" or "否：原因"
        account_match: "是" or "否：原因"
        follower_check: "是" or "否：原因"
        view_check: "是" or "否：原因"
        position_check: "是" or "否：原因"

    Returns:
        Validation result dictionary
    """
    return {
        "row_index": row_index,
        "validation_1": platform_match,
        "validation_2": title_match,
        "validation_3": time_match,
        "validation_4": account_match,
        "validation_5": follower_check,
        "validation_6": view_check,
        "validation_7": position_check,
    }


if __name__ == "__main__":
    # Example usage
    print("Excel utilities loaded successfully")
    print("\nAvailable functions:")
    print("  - read_excel_data(excel_path)")
    print("  - add_validation_columns(df, validation_results)")
    print("  - save_excel_with_validation(df, output_path, original_path)")
    print("  - create_validation_result(...)")
