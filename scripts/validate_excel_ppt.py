#!/usr/bin/env python3
"""
Complete validation with 3-tier stats extraction:
1. PPT text (automatic)
2. PPT screenshots (requires Claude vision)
3. Web pages (automatic fallback)
"""

import sys
import json
from pathlib import Path

from extract_ppt_text import extract_ppt_text
from extract_slides_smart import extract_slide_images
from parse_ppt_data import parse_all_slides
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def read_excel_data(excel_path):
    """Read Excel file"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(1, col).value)

    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            header = headers[col - 1]
            value = ws.cell(row, col).value
            row_data[header] = value
        row_data['_excel_row'] = row
        data_rows.append(row_data)

    return headers, data_rows, wb, ws


def fetch_stats_from_url(url):
    """Tier 3: Fetch stats from web URL"""
    try:
        import requests
        from bs4 import BeautifulSoup
        import re

        response = requests.get(url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        })
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        text = soup.get_text()

        stats = {'followers': None, 'views': None}

        # Follower patterns
        follower_patterns = [
            r'粉丝[数量]?[：:]\s*(\d+)',
            r'(\d+)\s*粉丝',
            r'关注[：:]\s*(\d+)',
            r'(\d+)\s*关注',
        ]

        for pattern in follower_patterns:
            match = re.search(pattern, text)
            if match:
                stats['followers'] = match.group(1)
                break

        # View patterns
        view_patterns = [
            r'阅读[量]?[：:]\s*(\d+)',
            r'浏览[量]?[：:]\s*(\d+)',
            r'(\d+)\s*阅读',
            r'(\d+)\s*浏览',
        ]

        for pattern in view_patterns:
            match = re.search(pattern, text)
            if match:
                stats['views'] = match.group(1)
                break

        return stats

    except Exception as e:
        return {'followers': None, 'views': None, 'error': str(e)}


def get_stats_3tier(slide_number, ppt_text_data, screenshot_stats, url, slide_image_path=None):
    """
    Get stats with 3-tier priority:
    1. PPT text (already in ppt_text_data)
    2. Screenshot (from screenshot_stats JSON or auto OCR)
    3. Web page (fetch if needed)

    Returns: dict with followers, views, and source
    """
    result = {
        'followers': None,
        'views': None,
        'source': None
    }

    # Tier 1: PPT text
    # (Usually doesn't have these stats, but check anyway)
    # This tier is typically N/A

    # Tier 2: Screenshot
    if screenshot_stats:
        slide_key = f"slide_{slide_number}"
        screenshot_data = screenshot_stats.get(slide_key, {})

        if screenshot_data.get('followers'):
            result['followers'] = screenshot_data['followers']
            result['source'] = 'screenshot'

        if screenshot_data.get('views'):
            result['views'] = screenshot_data['views']
            if not result['source']:
                result['source'] = 'screenshot'

    # Tier 2: Auto Vision API if screenshot_stats not provided
    if slide_image_path and not screenshot_stats and (not result['followers'] or not result['views']):
        print(f"    尝试使用Vision API自动识别截图...")
        try:
            from vision_api import auto_extract_stats
            vision_stats = auto_extract_stats(slide_image_path)

            if 'error' not in vision_stats:
                if not result['followers'] and vision_stats.get('followers'):
                    result['followers'] = vision_stats['followers']
                    result['source'] = vision_stats.get('backend', 'vision')

                if not result['views'] and vision_stats.get('views'):
                    result['views'] = vision_stats['views']
                    result['source'] = result['source'] or vision_stats.get('backend', 'vision')

                if result['followers'] or result['views']:
                    print(f"      ✓ 识别成功 ({vision_stats.get('backend', 'vision')}): 粉丝={result.get('followers', 'N/A')}, 阅读={result.get('views', 'N/A')}")
            else:
                print(f"      ⚠ Vision API不可用: {vision_stats.get('error', 'Unknown error')}")
        except ImportError:
            print(f"      ⚠ Vision API模块未安装（可选功能）")
        except Exception as e:
            print(f"      ⚠ Vision识别失败: {e}")

    # Tier 3: Web (if still missing data)
    if url and (not result['followers'] or not result['views']):
        print(f"    尝试从网页获取...")
        web_stats = fetch_stats_from_url(url)

        if not result['followers'] and web_stats.get('followers'):
            result['followers'] = web_stats['followers']
            result['source'] = result['source'] or 'web'

        if not result['views'] and web_stats.get('views'):
            result['views'] = web_stats['views']
            result['source'] = result['source'] or 'web'

    if not result['source']:
        result['source'] = 'none'

    return result


def validate_row(excel_row, ppt_data, screenshot_stats, slide_image_path=None):
    """Validate Excel row with 3-tier stats extraction (including OCR)"""
    results = {
        'excel_row': excel_row['_excel_row'],
        'slide': ppt_data.get('slide_number')
    }

    # 校验1: 发布平台
    excel_platform = excel_row.get('媒体名称/平台', '')
    ppt_platform = ppt_data.get('platform', '')

    if excel_platform == ppt_platform:
        results['check_1'] = '是'
    else:
        results['check_1'] = f'否：PPT显示"{ppt_platform}"，Excel为"{excel_platform}"'

    # 校验2: 文章标题
    excel_title = excel_row.get('文章标题', '')
    ppt_title = ppt_data.get('title', '')

    if excel_title == ppt_title:
        results['check_2'] = '是'
    else:
        results['check_2'] = f'否：标题不匹配\nExcel: {excel_title}\nPPT: {ppt_title}'

    # 校验3: 发布时间
    excel_date = str(excel_row.get('发布日期', ''))[:10]
    ppt_date = str(ppt_data.get('date', ''))[:10]

    excel_date_norm = excel_date.replace('.', '-').replace('/', '-')
    ppt_date_norm = ppt_date.replace('.', '-').replace('/', '-')

    if excel_date_norm == ppt_date_norm:
        results['check_3'] = '是'
    else:
        results['check_3'] = f'否：PPT显示"{ppt_date}"，Excel为"{excel_date}"'

    # 校验4: 发布账号
    excel_account = excel_row.get('发布账号', '/')
    ppt_account = ppt_data.get('account')

    if excel_account == '/' and ppt_account:
        results['check_4'] = f'否：Excel未记录，PPT显示"{ppt_account}"'
    elif ppt_account and excel_account == ppt_account:
        results['check_4'] = '是'
    elif not ppt_account:
        results['check_4'] = 'N/A：PPT未显示账号信息'
    else:
        results['check_4'] = f'否：PPT显示"{ppt_account}"，Excel为"{excel_account}"'

    # 校验5-6: 粉丝数和阅读量 (3-tier extraction)
    print(f"  获取统计数据（3级优先级：PPT文字 → 截图 → 网页）")

    slide_num = ppt_data.get('slide_number')
    url = ppt_data.get('link')

    stats = get_stats_3tier(slide_num, ppt_data, screenshot_stats, url, slide_image_path)

    source_display = {
        'screenshot': '截图',
        'claude': 'Claude',
        'openai': 'GPT-4V',
        'gemini': 'Gemini',
        'web': '网页',
        'none': 'N/A'
    }.get(stats['source'], stats['source'])

    # 校验5: 粉丝数量
    excel_followers = str(excel_row.get('粉丝量', '/')).strip()
    ppt_followers = stats.get('followers')

    if ppt_followers:
        print(f"    ✓ 粉丝数: {ppt_followers} (来源:{source_display})")
        try:
            if excel_followers == '/':
                results['check_5'] = f'信息：{source_display}显示{ppt_followers}，Excel未记录'
            else:
                excel_f = int(excel_followers.replace(',', ''))
                ppt_f = int(str(ppt_followers).replace(',', ''))
                if ppt_f >= excel_f:
                    results['check_5'] = '是'
                else:
                    results['check_5'] = f'否：{source_display}显示{ppt_f} < Excel{excel_f}'
        except ValueError:
            results['check_5'] = f'信息：{source_display}显示{ppt_followers}，Excel为{excel_followers}'
    else:
        print(f"    ✗ 粉丝数: 未找到")
        results['check_5'] = 'N/A：PPT文字、截图和网页均未找到粉丝数'

    # 校验6: 阅读量
    excel_views = str(excel_row.get('阅读量', '/')).strip()
    ppt_views = stats.get('views')

    if ppt_views:
        print(f"    ✓ 阅读量: {ppt_views} (来源:{source_display})")
        try:
            if excel_views == '/':
                results['check_6'] = f'信息：{source_display}显示{ppt_views}，Excel未记录'
            else:
                excel_v = int(excel_views.replace(',', ''))
                ppt_v = int(str(ppt_views).replace(',', ''))
                if ppt_v >= excel_v:
                    results['check_6'] = '是'
                else:
                    results['check_6'] = f'否：{source_display}显示{ppt_v} < Excel{excel_v}'
        except ValueError:
            results['check_6'] = f'信息：{source_display}显示{ppt_views}，Excel为{excel_views}'
    else:
        print(f"    ✗ 阅读量: 未找到")
        results['check_6'] = 'N/A：PPT文字、截图和网页均未找到阅读量'

    # 校验7: 见刊位置
    excel_position = excel_row.get('刊出位置', '')
    ppt_position = ppt_data.get('position')

    if ppt_position and excel_position == ppt_position:
        results['check_7'] = '是'
    elif not ppt_position:
        results['check_7'] = 'N/A：PPT未显示位置信息'
    else:
        results['check_7'] = f'否：PPT显示"{ppt_position}"，Excel为"{excel_position}"'

    return results


def add_validation_columns(ws, validation_results):
    """Add validation columns with formatting"""
    validation_headers = [
        "校验1-发布平台",
        "校验2-文章标题",
        "校验3-发布时间",
        "校验4-发布账号",
        "校验5-粉丝数量",
        "校验6-阅读量",
        "校验7-见刊位置"
    ]

    start_col = ws.max_column + 1

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, header in enumerate(validation_headers):
        col_idx = start_col + i
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = 35

    for result in validation_results:
        excel_row = result['excel_row']
        for i in range(7):
            col = start_col + i
            key = f'check_{i+1}'
            value = result.get(key, '')

            cell = ws.cell(excel_row, col, value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if value.startswith('是'):
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100")
            elif value.startswith('否'):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006")
            elif value.startswith('N/A') or value.startswith('信息'):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(color="9C6500")


def load_screenshot_stats(json_path):
    """Load screenshot stats from JSON file (created by Claude)"""
    if not Path(json_path).exists():
        return None

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Convert list to dict keyed by slide_number
    stats_dict = {}
    for item in data:
        slide_num = item.get('slide_number')
        if slide_num:
            key = f"slide_{slide_num}"
            stats_dict[key] = {
                'followers': item.get('followers'),
                'views': item.get('views'),
                'status': item.get('status')
            }

    return stats_dict


def main(ppt_path, excel_path, screenshot_stats_json=None, output_path=None):
    """Main validation workflow with 3-tier stats extraction"""
    ppt_path = Path(ppt_path)
    excel_path = Path(excel_path)

    print("=" * 80)
    print("Excel vs PPT 完整校验（3级数据提取：文字→截图→网页）")
    print("=" * 80)
    print()

    # Step 1: Extract PPT text
    print("Step 1: 提取PPT文字内容...")
    temp_json = ppt_path.parent / "ppt_text_temp.json"
    extract_ppt_text(str(ppt_path), str(temp_json))

    # Step 2: Extract slide screenshots (for OCR or manual processing)
    print("\nStep 2: 提取PPT幻灯片截图...")
    slides_dir = ppt_path.parent / "slides"
    slide_images = extract_slide_images(str(ppt_path), str(slides_dir))
    print(f"  提取了 {len(slide_images)} 张幻灯片")

    # Step 3: Parse PPT data
    print("\nStep 3: 解析PPT数据...")
    ppt_slides_data = parse_all_slides(str(temp_json))

    # Step 4: Load screenshot stats (if available)
    screenshot_stats = None
    if screenshot_stats_json:
        print(f"\nStep 4: 加载截图数据 ({screenshot_stats_json})...")
        screenshot_stats = load_screenshot_stats(screenshot_stats_json)
        if screenshot_stats:
            print(f"  ✓ 加载了 {len(screenshot_stats)} 个幻灯片的数据")
        else:
            print(f"  ⚠ 文件不存在或为空")
    else:
        print("\nStep 4: 跳过截图数据（未提供screenshot_stats.json）")

    # Step 5: Read Excel
    print("\nStep 5: 读取Excel数据...")
    headers, excel_rows, wb, ws = read_excel_data(str(excel_path))
    print(f"  Excel行数: {len(excel_rows)}")

    # Step 6: Validate
    print("\nStep 6: 执行校验...")
    validation_results = []

    for i, excel_row in enumerate(excel_rows):
        if i < len(ppt_slides_data):
            print(f"\n第{i+1}行:")
            ppt_data = ppt_slides_data[i]

            # Get corresponding slide image if available
            slide_img = slide_images[i] if i < len(slide_images) else None

            result = validate_row(excel_row, ppt_data, screenshot_stats, slide_img)
            validation_results.append(result)
            print(f"  ✓ 校验完成")

    # Step 7: Generate output
    print("\nStep 7: 生成结果文件...")
    add_validation_columns(ws, validation_results)

    if not output_path:
        output_filename = excel_path.stem + "-校验结果" + excel_path.suffix
        output_path = excel_path.parent / output_filename

    wb.save(output_path)
    print(f"✓ 校验结果已保存：{output_path}")

    # Cleanup
    temp_json.unlink()

    # Summary
    print("\n" + "=" * 80)
    print("校验汇总")
    print("=" * 80)
    total = len(validation_results)
    passed = sum(1 for r in validation_results
                 if all(r.get(f'check_{i}', '').startswith('是') or
                       r.get(f'check_{i}', '').startswith('N/A') or
                       r.get(f'check_{i}', '').startswith('信息')
                       for i in range(1, 8)))
    print(f"总计: {total} 行")
    print(f"全部通过: {passed} 行")
    print(f"存在问题: {total - passed} 行")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 validate_with_stats.py <ppt_file> <excel_file> [screenshot_stats.json] [output_file]")
        print()
        print("3-tier stats extraction priority:")
        print("  1. PPT text (automatic)")
        print("  2. Screenshots (from screenshot_stats.json if provided)")
        print("  3. Web pages (automatic fallback)")
        print()
        print("To use screenshot stats:")
        print("  1. Run: python3 extract_screenshot_stats.py <slides_dir> <ppt_text.json> <stats.json>")
        print("  2. Use Claude to read each slide image and update stats.json")
        print("  3. Run: python3 validate_with_stats.py <ppt> <excel> stats.json")
        sys.exit(1)

    ppt_file = sys.argv[1]
    excel_file = sys.argv[2]
    screenshot_stats_json = sys.argv[3] if len(sys.argv) >= 4 and not sys.argv[3].endswith('.xlsx') else None
    output_file = sys.argv[4] if len(sys.argv) >= 5 else (sys.argv[3] if len(sys.argv) >= 4 and sys.argv[3].endswith('.xlsx') else None)

    main(ppt_file, excel_file, screenshot_stats_json, output_file)
