#!/usr/bin/env python3
"""
全自动完整校验流程 —— 文字优先 + 图片兜底（支持 Kimi K2.5 / 豆包）
流程：
1. 提取PPT文字 → 保存到 txt
2. 提取每页所有图片 → 保存到文件夹
3. 文字优先搜索需要的字段（平台/标题/时间/账号/位置/粉丝/阅读）
   - 文字找到就用文字，文字找不到才用图片识别
4. 图片兜底识别缺失字段 → 调用项目已有的 extract_content (兼容 Kimi)
   → 每张图片都识别，**并发处理**提高速度，收集所有结果 → 粉丝/阅读分别取最大值
5. 解析PPT文字得到结构化数据
6. 和Excel对比校验
7. 生成带颜色标记的结果Excel

完全自动化，一步到位：
uv run scripts/full_auto_validate.py <ppt_file> <excel_file> [output_file]

支持环境变量配置：
-  Kimi K2.5:
    MOONSHOT_API_KEY  or  MOONSHOT_MODEL_KEY
    MOONSHOT_MODEL  (optional, default: kimi-k2.5)

-  OPENAI 兼容格式（豆包/Anthropic/OpenAI）:
    OPENCLAW_VISION_API_KEY
    OPENCLAW_VISION_BASE_URL  (optional, default: https://ark.cn-beijing.volces.com/api/coding/v3)
    OPENCLAW_VISION_MODEL   (optional, default: doubao-seed-2.0-code)

自动检测：如果找到 MOONSHOT_API_KEY 就用 Kimi，否则用 OPENCLAW_VISION。
"""

import sys
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from extract_ppt_text import extract_ppt_text
from extract_all_images import extract_all_images
from parse_ppt_data import parse_all_slides
from extract_content import extract as kimi_extract


# ── 工具函数 ──────────────────────────────────────────────────────────────

def normalize_number(value):
    """Normalize number from string to int for comparison"""
    if value is None or value == 'null' or value == 'None' or value == '' or value is None:
        return None
    
    if isinstance(value, (int, float)):
        return int(value)
    
    value = str(value).strip().lower()
    
    # Remove thousands separators: "41, 682" → "41682"
    value = value.replace(',', '').replace(' ', '')
    
    # Handle units
    if '万' in value:
        num_str = value.replace('万', '').strip()
        try:
            num = float(num_str)
            return int(num * 10000)
        except ValueError:
            return None
    if 'k' in value or '千' in value:
        num_str = value.replace('k', '').replace('千', '').strip()
        try:
            num = float(num_str)
            return int(num * 1000)
        except ValueError:
            return None
    
    # Plain number
    try:
        value = value.replace(',', '').replace(' ', '')
        return int(float(value))
    except ValueError:
        # Extract all digits if still fails
        import re
        digits = re.findall(r'\d+', value)
        if digits:
            try:
                return int(''.join(digits))
            except ValueError:
                pass
        return None


def get_client():
    """Get vision client based on environment variables"""
    # Check for Kimi first
    moonshot_key = os.environ.get('MOONSHOT_API_KEY') or os.environ.get('MOONSHOT_MODEL_KEY')
    if moonshot_key:
        os.environ['MOONSHOT_API_KEY'] = moonshot_key
        return 'kimi', None
    
    # Fallback to OpenCLaw vision (OpenAI compatible)
    openclaw_key = os.environ.get('OPENCLAW_VISION_API_KEY')
    if openclaw_key:
        from openai import OpenAI
        base_url = os.environ.get('OPENCLAW_VISION_BASE_URL', 'https://ark.cn-beijing.volces.com/api/coding/v3')
        model = os.environ.get('OPENCLAW_VISION_MODEL', 'doubao-seed-2.0-code')
        client = OpenAI(api_key=openclaw_key, base_url=base_url)
        return 'openai', (client, model)
    
    raise ValueError(
        "❌ 找不到API密钥，请设置环境变量:\n"
        "  - 对于 Kimi K2.5:  export MOONSHOT_API_KEY=your-key\n"
        "  - 对于 OpenAI 兼容（豆包/Anthropic/OpenAI）:  export OPENCLAW_VISION_API_KEY=your-key\n"
    )


def process_single_image_kimi(img_path: str, fields):
    """Process a single image with Kimi, for concurrent execution"""
    img_name = Path(img_path).name
    results = kimi_extract([img_path], fields)
    if not results:
        return {
            'followers': None,
            'views': None,
            'image': img_path,
            'img_name': img_name,
        }
    result = results[0]
    return {
        'followers': result.get('followers'),
        'views': result.get('views'),
        'image': img_path,
        'img_name': img_name,
    }


def process_slide_images_kimi(slide_images: list[str]):
    """
    Process all images in one slide using Kimi extract_content (already compatible)
    If multiple images have data, keep the MAXIMUM value (reading grows over time)
    **Concurrent processing** for speed
    """
    import concurrent.futures
    
    followers_candidates: list[tuple[int, str | int, str]] = []  # (normalized, original, image_path)
    views_candidates: list[tuple[int, str | int, str]] = []
    
    fields = [
        {"name": "followers", "description": "粉丝数/粉丝量/关注数量", "type": "number"},
        {"name": "views", "description": "阅读量/浏览量/点击数量", "type": "number"},
    ]
    
    print(f"      并发识别 {len(slide_images)} 张图片...")
    
    # Concurrent processing
    processed = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(slide_images), 8)) as executor:
        futures = [executor.submit(process_single_image_kimi, img, fields) for img in slide_images]
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            processed.append(result)
            
            img_name = result['img_name']
            img_path = result['image']
            followers = result['followers']
            views = result['views']
            
            if followers is not None and followers != 'null' and followers != '':
                norm_val = normalize_number(followers)
                if norm_val is not None:
                    followers_candidates.append((norm_val, followers, img_path))
                    print(f"      ✓ 找到粉丝数: {followers} (在 {img_name})")
            
            if views is not None and views != 'null' and views != '':
                norm_val = normalize_number(views)
                if norm_val is not None:
                    views_candidates.append((norm_val, views, img_path))
                    print(f"      ✓ 找到阅读量: {views} (在 {img_name})")
    
    # After processing all images, select MAX
    result = {
        'followers': None,
        'views': None,
        'processed': processed,
        'found_in': None,
    }
    
    if followers_candidates:
        followers_candidates.sort(key=lambda x: x[0])
        largest = followers_candidates[-1]
        result['followers'] = largest[1]
        result['found_in'] = largest[2]
        if len(followers_candidates) > 1:
            print(f"      ⚙ 选择最大值: {largest[1]} (共 {len(followers_candidates)} 个候选)")
    
    if views_candidates:
        views_candidates.sort(key=lambda x: x[0])
        largest = views_candidates[-1]
        result['views'] = largest[1]
        result['found_in'] = largest[2]
        if len(views_candidates) > 1:
            print(f"      ⚙ 选择最大值: {largest[1]} (共 {len(views_candidates)} 个候选)")
    
    if result['found_in'] is None:
        if followers_candidates:
            result['found_in'] = followers_candidates[-1][2]
        elif views_candidates:
            result['found_in'] = views_candidates[-1][2]
    
    return result


def extract_stats_from_image_openai(image_path: str, client, model: str):
    """Extract followers and views from image using OpenAI compatible vision"""
    import base64
    
    base64_image = base64.b64encode(open(image_path, "rb").read()).decode('utf-8')
    
    prompt = """
请仔细查看这张图片，提取以下数据：
1. 粉丝数/粉丝量/关注数 - 找出任何表示粉丝/关注数量的数字
2. 阅读量/浏览量/点击量 - 找出任何表示阅读/浏览/点击数量的数字

你只需要提取数字，可以带单位（如 27.6万, 1.3k）。
请只返回JSON格式，不要其他文字：
{
  "followers": "提取到的粉丝数（如果没找到就是null）",
  "views": "提取到的阅读量（如果没找到就是null）"
}
"""

    # Kimi requires temperature=1, others can use 0.0
    if 'kimi' in model.lower() or 'moonshot' in model.lower():
        temperature = 1.0
    else:
        temperature = 0.0

    response = client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        }
                    },
                ],
            }
        ],
        max_tokens=200,
        temperature=temperature,
    )
    
    result_text = response.choices[0].message.content.strip()
    
    # Clean up markdown code block
    if result_text.startswith('```json'):
        result_text = result_text[7:]
    if result_text.endswith('```'):
        result_text = result_text[:-3]
    result_text = result_text.strip()
    
    try:
        result = json.loads(result_text)
        return {
            'followers': result.get('followers'),
            'views': result.get('views'),
            'image': str(image_path),
            'img_name': Path(image_path).name,
        }
    except json.JSONDecodeError as e:
        return {
            'followers': None,
            'views': None,
            'image': str(image_path),
            'img_name': Path(image_path).name,
            'error': result_text[:100],
        }


def process_slide_images_openai(slide_images: list[str], client, model):
    """OpenAI compatible version — **Concurrent processing** for speed"""
    import concurrent.futures
    
    followers_candidates: list[tuple[int, str, str]] = []  # (normalized, original, image_path)
    views_candidates: list[tuple[int, str, str]] = []
    
    print(f"      并发识别 {len(slide_images)} 张图片...")
    
    # Concurrent processing
    processed = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(slide_images), 8)) as executor:
        futures = [executor.submit(extract_stats_from_image_openai, img, client, model) for img in slide_images]
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            processed.append(result)
            
            img_name = result['img_name']
            img_path = result['image']
            followers = result['followers']
            views = result['views']
            
            if followers and followers != 'null':
                norm_val = normalize_number(followers)
                if norm_val is not None:
                    followers_candidates.append((norm_val, followers, img_path))
                    print(f"      ✓ 找到粉丝数: {followers} (在 {img_name})")
            
            if views and views != 'null':
                norm_val = normalize_number(views)
                if norm_val is not None:
                    views_candidates.append((norm_val, views, img_path))
                    print(f"      ✓ 找到阅读量: {views} (在 {img_name})")
    
    # After processing all images, select MAX
    result = {
        'followers': None,
        'views': None,
        'processed': processed,
        'found_in': None,
    }
    
    if followers_candidates:
        followers_candidates.sort(key=lambda x: x[0])
        largest = followers_candidates[-1]
        result['followers'] = largest[1]
        result['found_in'] = largest[2]
        if len(followers_candidates) > 1:
            print(f"      ⚙ 选择最大值: {largest[1]} (共 {len(followers_candidates)} 个候选)")
    
    if views_candidates:
        views_candidates.sort(key=lambda x: x[0])
        largest = views_candidates[-1]
        result['views'] = largest[1]
        result['found_in'] = largest[2]
        if len(views_candidates) > 1:
            print(f"      ⚙ 选择最大值: {largest[1]} (共 {len(views_candidates)} 个候选)")
    
    if result['found_in'] is None:
        if followers_candidates:
            result['found_in'] = followers_candidates[-1][2]
        elif views_candidates:
            result['found_in'] = views_candidates[-1][2]
    
    return result


def read_excel_data(excel_path: Path):
    """Read Excel data"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(1, col).value)
    
    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            header = headers[col - 1]
            value = ws.cell(row, col).value
            row_data[header] = value
        row_data['_excel_row'] = row
        data_rows.append(row_data)
    
    return headers, data_rows, wb, ws


def validate_row(excel_row: dict, ppt_data: dict, slide_image_stats: dict):
    """Validate one row against Excel"""
    results = {
        'excel_row': excel_row['_excel_row'],
        'slide': ppt_data.get('slide_number'),
    }

    # 校验1: 发布平台
    excel_platform = excel_row.get('媒体名称/平台', '')
    ppt_platform = ppt_data.get('platform', '')

    if excel_platform == ppt_platform:
        results['check_1'] = '是'
    else:
        results['check_1'] = f'否：PPT显示"{ppt_platform}"，Excel为"{excel_platform}"'

    # 校验2: 文章标题
    excel_title = excel_row.get('文章标题', '')
    ppt_title = ppt_data.get('title', '')

    if excel_title == ppt_title:
        results['check_2'] = '是'
    else:
        results['check_2'] = f'否：标题不匹配\nExcel: {excel_title}\nPPT: {ppt_title}'

    # 校验3: 发布时间
    excel_date = str(excel_row.get('发布日期', ''))[:10]
    ppt_date = str(ppt_data.get('date', ''))[:10]

    excel_date_norm = excel_date.replace('.', '-').replace('/', '-')
    ppt_date_norm = ppt_date.replace('.', '-').replace('/', '-')

    if excel_date_norm == ppt_date_norm:
        results['check_3'] = '是'
    else:
        results['check_3'] = f'否：PPT显示"{ppt_date}"，Excel为"{excel_date}"'

    # 校验4: 发布账号
    excel_account = excel_row.get('发布账号', '/')
    ppt_account = ppt_data.get('account')

    # 规则:
    # - Excel是/ (无)，PPT也无 → N/A
    # - Excel是/ (无)，PPT有 → 否: Excel未记录，PPT有
    # - Excel有，PPT有，相等 → 是
    # - Excel有，PPT有，不等 → 否
    # - Excel有，PPT无 → N/A

    if (excel_account == '/' or excel_account == '') and not ppt_account:
        results['check_4'] = 'N/A：Excel和PPT都未记录账号信息'
    elif (excel_account == '/' or excel_account == '') and ppt_account:
        results['check_4'] = f'否：Excel未记录，PPT显示"{ppt_account}"'
    elif ppt_account and str(excel_account).strip() == str(ppt_account).strip():
        results['check_4'] = '是'
    elif not ppt_account:
        results['check_4'] = 'N/A：PPT未显示账号信息'
    else:
        results['check_4'] = f'否：PPT显示"{ppt_account}"，Excel为"{excel_account}"'

    # 校验5-6: 粉丝数和阅读量 (来自图片识别，取最大值)
    print(f"  获取统计数据（文字优先，图片兜底，收集结果取最大值）")
    
    followers = slide_image_stats.get('followers')
    views = slide_image_stats.get('views')
    found_in = slide_image_stats.get('found_in')
    
    source_display = f"自动识别 ({Path(found_in).parent.name}/{Path(found_in).name})" if found_in else "自动识别"

    # 校验5: 粉丝数量
    excel_followers = str(excel_row.get('粉丝量', '/')).strip()
    followers_norm = normalize_number(followers)
    
    if followers and followers != 'null':
        print(f"    ✓ 粉丝数: {followers} (来源:{source_display})")
        try:
            if excel_followers == '/' or excel_followers == 'nan' or excel_followers == '':
                results['check_5'] = f'信息：{source_display}显示{followers}，Excel未记录'
            else:
                excel_f = normalize_number(excel_followers)
                ppt_f = followers_norm
                if excel_f is None or ppt_f >= excel_f:
                    results['check_5'] = '是'
                else:
                    results['check_5'] = f'否：{source_display}显示{ppt_f} < Excel{excel_f}'
        except ValueError:
            results['check_5'] = f'信息：{source_display}显示{followers}，Excel为{excel_followers}'
    else:
        print(f"    ✗ 粉丝数: 所有图片中都未找到")
        results['check_5'] = 'N/A：这页PPT所有图片中均未找到粉丝数'

    # 校验6: 阅读量
    excel_views = str(excel_row.get('阅读量', '/')).strip()
    views_norm = normalize_number(views)
    
    if views and views != 'null':
        print(f"    ✓ 阅读量: {views} (来源:{source_display})")
        try:
            if excel_views == '/' or excel_views == 'nan' or excel_views == '':
                results['check_6'] = f'信息：{source_display}显示{views}，Excel未记录'
            else:
                excel_v = normalize_number(excel_views)
                ppt_v = views_norm
                if excel_v is None or ppt_v >= excel_v:
                    results['check_6'] = '是'
                else:
                    results['check_6'] = f'否：{source_display}显示{ppt_v} < Excel{excel_v}'
        except ValueError:
            results['check_6'] = f'信息：{source_display}显示{views}，Excel为{excel_views}'
    else:
        print(f"    ✗ 阅读量: 所有图片中都未找到")
        results['check_6'] = 'N/A：这页PPT所有图片中均未找到阅读量'

    # 校验7: 见刊位置
    excel_position = excel_row.get('刊出位置', '')
    ppt_position = ppt_data.get('position')

    if ppt_position and excel_position == ppt_position:
        results['check_7'] = '是'
    elif not ppt_position:
        results['check_7'] = 'N/A：PPT未显示位置信息'
    else:
        results['check_7'] = f'否：PPT显示"{ppt_position}"，Excel为"{excel_position}"'

    return results


def add_validation_columns(ws, validation_results):
    """Add validation columns to Excel with coloring"""
    validation_headers = [
        "校验1-发布平台",
        "校验2-文章标题",
        "校验3-发布时间",
        "校验4-发布账号",
        "校验5-粉丝数量",
        "校验6-阅读量",
        "校验7-见刊位置"
    ]

    start_col = ws.max_column + 1

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, header in enumerate(validation_headers):
        col_idx = start_col + i
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = 40

    for result in validation_results:
        excel_row = result['excel_row']
        for i in range(7):
            col = start_col + i
            key = f'check_{i+1}'
            value = result.get(key, '')

            cell = ws.cell(excel_row, col, value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if value.startswith('是'):
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100")
            elif value.startswith('否'):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006")
            elif value.startswith('N/A') or value.startswith('信息'):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(color="9C6500")


def main(ppt_path: str, excel_path: str, output_path: str = None):
    """Main workflow"""
    # Get client
    try:
        client_type, client_data = get_client()
    except ValueError as e:
        print(e)
        sys.exit(1)
    
    ppt_path = Path(ppt_path)
    excel_path = Path(excel_path)

    print("=" * 80)
    print("Batch Validator - 全自动完整校验")
    print("流程: 提取文字 → 提取所有图片 → 文字优先搜索 → 图片兜底识别缺失字段 → 汇总取最大值 → 校验 → 输出")
    print(f"Vision 模型: {'Kimi K2.5' if client_type == 'kimi' else 'OpenAI 兼容'}")
    print("并发处理: ✅ 开启 (最大 8 并发)")
    print("=" * 80)
    print()

    # Step 1: Extract PPT text
    print("Step 1: 提取PPT文字内容...")
    temp_json = ppt_path.parent / "ppt_text_full_auto.json"
    extract_ppt_text(str(ppt_path), str(temp_json))
    print()

    # Step 2: Extract all images from all slides
    print("Step 2: 提取每页所有图片...")
    images_dir = ppt_path.parent / f"{ppt_path.stem}_all_images"
    slide_images_dict = extract_all_images(str(ppt_path), str(images_dir))
    total_images = sum(len(imgs) for imgs in slide_images_dict.values())
    print(f"  总计 {len(slide_images_dict)} 页, {total_images} 张图片")
    print()

    # Step 3: Auto-recognize missing stats from all images
    print("Step 3: 图片识别提取统计数据 (每张图片都识别，最后取最大值)...")
    print()
    
    auto_stats = {}
    
    for slide_num, slide_images in slide_images_dict.items():
        if not slide_images:
            print(f"  第 {slide_num} 页: 没有图片，跳过")
            auto_stats[f"slide_{slide_num}"] = {
                'followers': None,
                'views': None,
                'processed': [],
                'found_in': None
            }
            continue
        
        print(f"  第 {slide_num} 页: {len(slide_images)} 张图片")
        if client_type == 'kimi':
            result = process_slide_images_kimi(slide_images)
        else:
            client, model = client_data
            result = process_slide_images_openai(slide_images, client, model)
        auto_stats[f"slide_{slide_num}"] = result
        print()
    
    # Save auto stats to JSON
    auto_stats_json = ppt_path.parent / "auto_full_stats.json"
    with open(auto_stats_json, 'w', encoding='utf-8') as f:
        json.dump(auto_stats, f, indent=2, ensure_ascii=False)
    print(f"✓ 自动识别结果已保存: {auto_stats_json}")
    print()

    # Step 4: Parse PPT data
    print("Step 4: 解析PPT数据...")
    ppt_slides_data = parse_all_slides(str(temp_json))
    print(f"  解析完成，得到 {len(ppt_slides_data)} 页数据")
    print()

    # Step 5: Read Excel
    print("Step 5: 读取Excel数据...")
    headers, excel_rows, wb, ws = read_excel_data(excel_path)
    print(f"  Excel行数: {len(excel_rows)}")
    print()

    # Step 6: Validate
    print("Step 6: 执行校验...")
    validation_results = []

    for i, excel_row in enumerate(excel_rows):
        if i < len(ppt_slides_data):
            print(f"\n第{i+1}行:")
            ppt_data = ppt_slides_data[i]
            slide_num = ppt_data.get('slide_number')
            slide_image_stats = auto_stats.get(f"slide_{slide_num}", {})
            
            result = validate_row(excel_row, ppt_data, slide_image_stats)
            validation_results.append(result)
            print(f"  ✓ 校验完成")

    # Step 7: Generate output
    print("\nStep 7: 生成结果文件...")
    add_validation_columns(ws, validation_results)

    if not output_path:
        output_filename = excel_path.stem + "-全自动校验结果" + excel_path.suffix
        output_path = excel_path.parent / output_filename

    wb.save(output_path)
    print(f"✓ 校验结果已保存：{output_path}")

    # Summary
    print("\n" + "=" * 80)
    print("校验汇总")
    print("=" * 80)
    total = len(validation_results)
    passed = sum(1 for r in validation_results
                 if all(r.get(f'check_{i}', '').startswith('是') or
                       r.get(f'check_{i}', '').startswith('N/A') or
                       r.get(f'check_{i}', '').startswith('信息')
                       for i in range(1, 8)))
    print(f"总计: {total} 行")
    print(f"全部通过: {passed} 行")
    print(f"存在问题: {total - passed} 行")
    print(f"\n图片识别结果: {auto_stats_json}")
    print(f"输出文件: {output_path}")
    
    # List issues
    if total - passed > 0:
        print(f"\n存在问题的行:")
        for r in validation_results:
            issues = [r[f'check_{i}'] for i in range(1, 8) if r[f'check_{i}'].startswith('否')]
            if issues:
                print(f"  第{r['excel_row']}行: {issues[0]}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: uv run scripts/full_auto_validate.py <ppt文件> <excel文件> [输出文件]")
        print()
        print("✨ 全自动流程:")
        print("  1. 提取PPT文字")
        print("  2. 提取每页所有图片（多张不漏）")
        print("  3. 文字优先查询结构化信息，文字找不到才图片识别")
        print("  4. 每张图片都识别，**并发处理**提高速度，收集所有找到的值 → 粉丝/阅读分别取最大值")
        print("    → 符合需求：阅读量随着时间会快速变化，我们应该取最大的那个值")
        print("  5. 和Excel对比校验")
        print("  6. 生成带颜色标记的结果Excel")
        print()
        print("环境变量配置:")
        print("  - Kimi K2.5:  export MOONSHOT_API_KEY=your-key")
        print("  - OpenAI兼容（豆包）:  export OPENCLAW_VISION_API_KEY=your-key")
        print()
        sys.exit(1)

    ppt_file = sys.argv[1]
    excel_file = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) >= 4 else None

    main(ppt_file, excel_file, output_file)
